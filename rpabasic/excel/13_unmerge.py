# 셀 병합
from openpyxl import load_workbook

wb = load_workbook("./rpabasic/excel/merge.xlsx")
ws = wb.active

# 병합 해제 범위 지정
ws.unmerge_cells("B2:D2")


wb.save("./rpabasic/excel/unmerge.xlsx")
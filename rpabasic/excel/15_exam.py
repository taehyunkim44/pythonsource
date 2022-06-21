from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])

scores =[
    [1,10,8,5,14,26,12],
    [2,7,3,7,15,24,18],
    [3,9,5,8,8,12,4],
    [4,7,8,7,17,21,18]
    [5,7,8,7,16,25,15]
    [6,3,5,8,8,17,0],
    [7,4,9,10,16,27,18],
    [8,6,6,6,15,19,17],
    [9,10,10,9,19,30,19],
    [10,9,8,8,20,25,20]
]

for score in scores:
    ws.append(score)

# 퀴즈2 점수 10 점으로 변경
for idx, cell in enumerate(ws["D"]):
    if idx == 0: # 제목행인 경우 skip
        continue
    cell.value = 10

# 총점, 성적 셀 추가
ws["H1"].value = "총점"
ws["I1"].value = "성적"

for idx,score in enumerate(scores,start=2):
    ws.cell(row=idx, column=8).value = "=sum(B{0}:G{0})".format(idx)

    total = sum(score[1:]) - score[3] + 10
    grade = None
    if total >= 90:
        grade = "A"
    elif total >= 80:
        grade = "B"
    elif total >= 70:
        grade = "C"
    else: 
        grade = "D"
    
    if score[1] < 5:
        grade = "F"

    ws.cell(row=idx, column=9).value = grade

wb.save("./rpabasic/excel/scores.xlsx")
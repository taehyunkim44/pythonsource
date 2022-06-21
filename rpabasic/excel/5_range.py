import random
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 한 줄씩 데이터 삽입 - 리스트 이용
ws.append(["번호", "영어", "수학"])  # 첫줄 입력
for i in range(1, 11):
    ws.append([i, random.randint(0, 100), random.randint(0, 100)])

# rows = [
#     ["이름", "생년월일"],
#     ["홍길동", "801020"],
#     ["성춘향", "811020"],
#     ["김지원", "860920"],
#     ["남주혁", "880705"],
# ]

# for row in rows:
#     ws.append(row)

# 첫번째 행 가져오기
print(ws["A1"].value, ws["B1"].value, ws["C1"].value)

for j in range(1, 4):
    print(ws.cell(i, j).value, end=" ")

first_row = ws[1]
for cell in first_row:
    print(cell.value, end=" ")

print()
# 2~6번 행 가져오기
row_range = ws[2:6]
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

# 3번행부터 마지막 까지 가져오기
row_range = ws[3 : ws.max_row]
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

# iter_rows() / iter_cols() : 전체 rows, cols 함수
for row in ws.iter_rows():
    print(row[2].value)

print()
for col in ws.iter_cols():
    print(col[2].value)


wb.save("./rpabasic/excel/range.xlsx")
